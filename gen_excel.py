import argparse
import os
import sys
import openpyxl

def main(argv):
    parser = argparse.ArgumentParser(description="Generate Excel File")
    parser.add_argument("--output", "-o", type=str, default="./superstars",
                        help="Output directory to save downloaded images.")
    parser.add_argument("--filename", "-f", type=str, default="output.xlsx",
                        help="Name of the Excel file.")
    args = parser.parse_args(args=argv)

    # 创建一个新的Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active

    # 设置首行的标题
    ws.cell(row=1, column=1, value="姓名")
    ws.cell(row=1, column=2, value="照片数")

    # 男明星
    # names = [
    #     "陈赫", "陈坤", "邓超", "冯绍峰", "龚俊", "胡歌", "黄渤", "黄磊", "黄晓明", "黄子韬", "贾乃亮", "李晨", "李现",
    #     "鹿晗","靳东", "井柏然", "雷佳音", "刘烨", "陆毅", "沈腾", "孙红雷", "佟大为", "王宝强", "汪峰", "王俊凯", "王凯",
    #     "王一博","王源", "吴京", "肖战", "徐峥", "杨洋", "易烊千玺", "岳云鹏", "张翰", "张杰", "张一山", "张艺兴", "赵文卓",
    #     "郑恺","朱一龙", "艾伦", "白敬亭", "白举纲", "白客", "白宇", "包贝尔", "陈若轩", "陈翔", "陈晓", "陈学冬", "成毅",
    #     "丁禹兮","董成鹏", "董子健", "窦骁", "杜淳", "杜江", "段奕宏", "高伟光", "郭京飞", "郭俊辰", "郭麒麟", "郭晓东", "韩栋",
    #     "韩东君","韩庚", "侯明昊", "胡兵", "胡一天", "黄景瑜", "黄轩", "黄宥明", "金瀚", "李宏毅", "李佳航", "林更新", "刘昊然",
    #     "刘宇宁","罗晋", "罗云熙", "马可", "马天宇", "毛不易", "茅子俊", "聂远", "欧豪", "潘粤明", "彭冠英", "彭昱畅", "乔振宇",
    #     "秦俊杰","任嘉伦", "任重", "沙溢", "盛一伦", "释小龙", "宋威龙", "孙坚", "孙艺洲", "檀健次", "王传君", "王鹤棣",
    #     "王千源", "王学兵","王迅", "王栎鑫", "魏晨", "魏大勋", "吴磊", "小沈阳", "肖央", "邢昭林", "徐海乔", "许凯", "徐开骋", "许魏洲",
    #     "严屹宽","杨烁", "杨旭文", "叶祖新", "印小天", "尹正", "俞灏明", "于和伟", "于朦胧", "袁弘", "曾舜晞", "张彬彬",
    #     "张嘉益", "张晋","张伦硕", "张睿", "张若昀", "张晓龙", "张新成", "张译", "张云龙", "周一围", "朱亚文"
    # ]
    # 女明星
    # names = [
    #     "Angelababy", "迪丽热巴", "关晓彤", "贾玲", "李冰冰", "刘诗诗", "刘涛", "刘亦菲", "倪妮", "宋佳", "宋茜",
    #     "孙俪", "唐嫣", "佟丽娅", "谢娜", "杨幂", "杨紫", "姚晨", "章子怡", "赵丽颖", "赵露思", "周冬雨", "周迅",
    #     "白百何", "白鹿", "蔡文静", "陈数", "陈钰琪", "陈紫函", "邓家佳", "董璇", "高露", "高圆圆", "古力娜扎",
    #     "海陆", "海清", "韩雪", "郝蕾", "胡冰卿", "胡静", "黄圣依", "黄奕", "霍思燕", "贾青", "江疏影", "蒋梦婕",
    #     "蒋勤勤", "蒋欣", "蒋依依", "江映蓉", "江一燕", "焦俊艳", "金晨", "景甜", "鞠婧祎", "阚清子", "辣目洋子",
    #     "李菲儿", "李沁", "李晟", "李溪芮", "李小冉", "李一桐", "林允", "刘浩存", "刘惜君", "柳岩", "刘芸", "娄艺潇",
    #     "吕一", "马丽", "马思纯", "马苏", "马伊俐", "毛晓彤", "梅婷", "孟美岐", "戚薇", "秦海璐", "秦岚", "任素汐",
    #     "沈梦辰", "沈月", "舒畅", "宋轶", "宋祖儿", "孙怡", "谭松韵", "谭维维", "唐艺昕", "童瑶", "万茜", "王丽坤",
    #     "王珞丹", "王鸥", "王晓晨", "王子文", "吴谨言", "吴倩", "吴宣仪", "邢菲", "徐璐", "杨超越", "杨蓉", "叶璇",
    #     "殷桃", "颖儿", "虞书欣", "袁冰妍", "袁泉", "袁姗姗", "赵今麦", "赵樱子", "张慧雯", "张俪", "张萌", "张含韵",
    #     "张嘉倪", "张天爱", "张小斐", "张歆艺", "张馨予", "张雪迎", "张雨绮", "张子枫", "左小青"
    # ]
    names = [
        "陈赫", "陈坤", "邓超"
    ]

    # 将人名写入Excel表格
    for idx, name in enumerate(names, start=2):
        ws.cell(row=idx, column=1, value=name)

    # 拼接文件的完整路径
    file_path = os.path.join(args.output, args.filename)

    # 确保输出目录存在
    os.makedirs(args.output, exist_ok=True)

    # 保存Excel文件
    wb.save(file_path)

    print(f"Excel文件已保存到指定目录下: {file_path}")


if __name__ == '__main__':
    main(sys.argv[1:])

# python gen_excel.py -f filename.xlsx